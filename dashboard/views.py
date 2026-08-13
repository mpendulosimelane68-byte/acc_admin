from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import render, get_object_or_404, redirect
from django.utils import timezone
from django.db.models import Count
from django.db.models.functions import TruncMonth
from django.db.models import Q
from dashboard.models import Case
from datetime import datetime
import requests
from .models import Notification
import json
from django.contrib.auth import logout
from django.shortcuts import redirect

from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.shortcuts import render, redirect
from .models import AuditLog
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.models import User
from django.contrib.auth.views import PasswordResetCompleteView
from .models import AuditLog
from django.contrib.auth.views import PasswordResetConfirmView
from .models import AuditLog
from django.http import HttpResponse
import openpyxl
from openpyxl import Workbook
from django.http import HttpResponse

from django.db.models import Count

from django.db.models.functions import TruncMonth

from .models import AuditLog, CaseStatusHistory, Notification
from .forms import CaseForm



# ---------------- DASHBOARD ----------------

from django.contrib.auth.decorators import login_required


@login_required(login_url="dashboard_login")
def dashboard(request):

    

    # Total cases
    total_cases = Case.objects.count()


    # Cases received today
    today = timezone.now().date()

    new_today = Case.objects.filter(
        created_at__date=today
    ).count()


    # Cases under investigation
    under_investigation = Case.objects.filter(
        status="investigation"
    ).count()


    # Cases referred elsewhere
    referrals = Case.objects.filter(
        status="referral"
    ).count()


    # Closed cases
    closed = Case.objects.filter(
        status="closed"
    ).count()


    # Recent cases
    recent_cases = Case.objects.all().order_by('-id')[:5]

    
    # Activity Feed  
    activity_logs = AuditLog.objects.all().order_by('-created_at')[:10]

    upcoming_notifications = Notification.objects.filter(
    date__gte=timezone.now().date()
).order_by('date', 'time')[:5]

       # WEATHER
    weather = {}

    try:
        response = requests.get(
            "https://wttr.in/Mbabane?format=j1"
        )

        data = response.json()

        weather = {
            "temperature": data["current_condition"][0]["temp_C"],
            "description": data["current_condition"][0]["weatherDesc"][0]["value"],
            "location": "Mbabane, Eswatini",
            "current_date": datetime.now().strftime("%A, %d %B %Y"),
        }

    except Exception:

        weather = {
            "temperature": "--",
            "description": "Unavailable",
            "location": "Mbabane, Eswatini",
            "current_date": datetime.now().strftime("%A, %d %B %Y"),
        }

# ==============================
# REPORTS MONTHLY TREND GRAPH
# ==============================

    monthly_cases = (
    Case.objects
    .annotate(month=TruncMonth('created_at'))
    .values('month')
    .annotate(total=Count('id'))
    .order_by('month')
)
   
    # ==============================
    # REPORTS MONTHLY TREND GRAPH
    # ==============================

    monthly_cases = (
        Case.objects
        .annotate(month=TruncMonth('created_at'))
        .values('month')
        .annotate(total=Count('id'))
        .order_by('month')
    )


    monthly_labels = [
        item['month'].strftime("%B")
        for item in monthly_cases
    ]


    monthly_values = [
        item['total']
        for item in monthly_cases
    ]



    # ==============================
    # PUBLIC VS PRIVATE GRAPH
    # ==============================

    sector_cases = (
        Case.objects
        .values('sector')
        .annotate(total=Count('id'))
    )


    sector_labels = [
        item['sector']
        for item in sector_cases
    ]


    sector_values = [
        item['total']
        for item in sector_cases
    ]


    # ==============================
    # STATUS GRAPH (optional)
    # ==============================

    status_cases = (
        Case.objects
        .values('status')
        .annotate(total=Count('id'))
    )


    status_labels = [
        item['status']
        for item in status_cases
    ]


    status_values = [
        item['total']
        for item in status_cases
    ]

    return render(
        request,
    "dashboard/dashboard.html",
    {

        # Cards
        "total_cases": total_cases,

        "new_today": new_today,

        "under_investigation": under_investigation,

        "referrals": referrals,

        "closed": closed,


        # Recent cases table
        "recent_cases": recent_cases,

       
        # Activity Feed
        "activity_logs": activity_logs,
        "upcoming_notifications": upcoming_notifications,
        "weather": weather,

        

         "current_date": datetime.now().strftime("%A, %d %B %Y"),
         "current_time": datetime.now().strftime("%H:%M"),


        # Monthly graph
        "monthly_labels": monthly_labels,

        "monthly_values": monthly_values,


        # Sector graph
        "sector_labels": sector_labels,

        "sector_values": sector_values,


        # Status graph
        "status_labels": status_labels,

        "status_values": status_values,

    }
)

def dashboard_login(request):

    if request.method == "POST":

        username = request.POST.get("username")
        password = request.POST.get("password")


        user = authenticate(
            request,
            username=username,
            password=password
        )


        if user is not None:

            login(request, user)

            AuditLog.objects.create(
                user=user,
                action="Login",
                description="Administrator logged into the system",
                ip_address=request.META.get("REMOTE_ADDR")
            )

            return redirect("dashboard")


        else:

            AuditLog.objects.create(
                user=None,
                action="Failed Login",
                description=f"Failed login attempt for username: {username}",
                ip_address=request.META.get("REMOTE_ADDR")
            )


            messages.error(
                request,
                "Invalid username or password"
            )


    return render(
        request,
        "dashboard/login.html"
    )

def dashboard_logout(request):

    if request.user.is_authenticated:

        AuditLog.objects.create(
            user=request.user,
            action="Logout",
            description="Administrator logged out of the system",
            ip_address=request.META.get("REMOTE_ADDR")
        )

    logout(request)

    response = redirect("dashboard_login")

    response["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response["Pragma"] = "no-cache"
    response["Expires"] = "0"

    return response
    
# ---------------- CASE LIST ----------------

def cases(request):

    cases_list = Case.objects.all().order_by("-id")

    # Status Cards
    total_cases = Case.objects.count()

    received_cases = Case.objects.filter(
    status="received"
).count()

    investigation_cases = Case.objects.filter(
    status="investigation"
).count()

    referral_cases = Case.objects.filter(
    status="referral"
).count()

    declined_cases = Case.objects.filter(
    status="declined"
).count()

    closed_cases = Case.objects.filter(
    status="closed"
).count()


    sector = request.GET.get("sector")
    region = request.GET.get("region")
    status = request.GET.get("status")
    gender = request.GET.get("gender")
    age_group = request.GET.get("age_group")
    search = request.GET.get("search")


    if sector:
        cases_list = cases_list.filter(
            sector=sector
        )


    if region:
        cases_list = cases_list.filter(
            region__iexact=region
        )


    if status:
        cases_list = cases_list.filter(
            status=status
        )


    if gender:
        cases_list = cases_list.filter(
            gender=gender
        )


    if age_group:
        cases_list = cases_list.filter(
            age_group=age_group
        )

    if search:
        cases_list = cases_list.filter(
            case_code__icontains=search
    )

    

    return render(
    request,
    "dashboard/cases.html",
    {
        "cases": cases_list,
        "search": search,

        "total_cases": total_cases,
        "received_cases": received_cases,
        "investigation_cases": investigation_cases,
        "referral_cases": referral_cases,
        "declined_cases": declined_cases,
        "closed_cases": closed_cases,
    }
)

def export_cases_excel(request):

    cases_list = Case.objects.all().order_by("-id")


    # Same filters as cases page

    sector = request.GET.get("sector")
    region = request.GET.get("region")
    status = request.GET.get("status")
    gender = request.GET.get("gender")
    age_group = request.GET.get("age_group")
    search = request.GET.get("search")


    if sector:
        cases_list = cases_list.filter(
            sector=sector
        )


    if region:
        cases_list = cases_list.filter(
            region__iexact=region
        )


    if status:
        cases_list = cases_list.filter(
            status=status
        )


    if gender:
        cases_list = cases_list.filter(
            gender=gender
        )


    if age_group:
        cases_list = cases_list.filter(
            age_group=age_group
        )


    if search:
        cases_list = cases_list.filter(
            case_code__icontains=search
        )


    workbook = openpyxl.Workbook()

    sheet = workbook.active

    sheet.title = "ACC Cases"


    sheet.append([
        "Case Code",
        "Anonymous",
        "Name",
        "Surname",
        "Contact",
        "Region",
        "Location",
        "Sector",
        "Status",
        "institution",
        "Amount",
        "Description",
        "Date Reported"
    ])


    for case in cases_list:

        sheet.append([
            case.case_code,
            "Yes" if case.is_anonymous else "No",
            case.name,
            case.surname,
            case.contact,
            case.region,
            case.location,
            case.sector,
            case.status,
            case.institution,
            case.amount,
            case.description,
            case.created_at.strftime("%Y-%m-%d")
        ])


    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


    response["Content-Disposition"] = (
        'attachment; filename="ACC_Cases.xlsx"'
    )


    workbook.save(response)


    return response



# ---------------- AUDIT LOGS ----------------

from django.utils import timezone
from django.shortcuts import render
from .models import AuditLog


@csrf_exempt
def get_case_status(request, case_code):

    if request.method != "GET":
        return JsonResponse(
            {
                "success": False,
                "error": "Only GET requests are allowed"
            },
            status=405
        )

    try:
        case = Case.objects.get(case_code=case_code)

        return JsonResponse(
            {
                "success": True,
                "case_code": case.case_code,
                "status": case.status,
                "decline_reason": case.decline_reason or "",
                "referral_notes": case.referral_notes or "",
                "description": case.description or "",
                "created_at": case.created_at.isoformat(),
            },
            status=200
        )

    except Case.DoesNotExist:
        return JsonResponse(
            {
                "success": False,
                "error": "Case not found"
            },
            status=404
        )

    except Exception as e:
        return JsonResponse(
            {
                "success": False,
                "error": str(e)
            },
            status=500
        )
    
def audit_logs(request):

    logs = AuditLog.objects.all().order_by('-id')

    # Activity Cards
    total_activities = AuditLog.objects.count()

    today = timezone.now().date()

    new_today = AuditLog.objects.filter(
        created_at__date=today
    ).count()

    for log in logs:

        # Case Code
        if log.case:
            log.display_case_code = log.case.case_code
        else:
            log.display_case_code = "-"


        # Action
        log.display_action = log.action


        # Description
        if log.description:
            log.display_description = log.description
        else:
            log.display_description = "No description"


        # User
        if hasattr(log, "user") and log.user:
            log.display_user = log.user.username
        else:
            log.display_user = "Unknown"


        # IP Address
        if log.ip_address:
            log.display_ip = log.ip_address
        else:
            log.display_ip = "Unknown"


    return render(
    request,
    "dashboard/audit_logs.html",
    {
        "logs": logs,
        "total_activities": total_activities,
        "new_today": new_today,
    }
)
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from .models import Case, AnonymousAccount


import random
import string

from django.db import IntegrityError
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt

from .models import Case, AnonymousAccount


def generate_unique_case_code():
    characters = string.ascii_uppercase + string.digits

    while True:
        part1 = ''.join(
            random.choices(characters, k=3)
        )
        part2 = ''.join(
            random.choices(characters, k=3)
        )

        code = f"ACC-{part1}-{part2}"

        if not Case.objects.filter(case_code=code).exists():
            return code


@csrf_exempt
def receive_report(request):

    if request.method != "POST":
        return JsonResponse(
            {
                "success": False,
                "error": "Only POST requests are allowed"
            },
            status=405
        )

    try:
        data = request.POST

        is_anonymous = (
            str(data.get("is_anonymous", "false")).lower() == "true"
        )

        # ------------------------------------------------------------
        # GET THE ACCOUNT THAT SUBMITTED THE REPORT
        # ------------------------------------------------------------

        username = data.get("username", "").strip()

        if not username:
            return JsonResponse(
                {
                    "success": False,
                    "error": "Username is required"
                },
                status=400
            )

        try:
            account = AnonymousAccount.objects.get(
                username=username
            )

        except AnonymousAccount.DoesNotExist:
            return JsonResponse(
                {
                    "success": False,
                    "error": "Account not found"
                },
                status=401
            )

        # ------------------------------------------------------------
        # CREATE CASE
        # ------------------------------------------------------------

        case = Case.objects.create(

            # Account identifies who submitted the report.
            # It does NOT determine whether the report is anonymous.
            account=account,

            case_code=generate_unique_case_code(),

            # This describes THIS report only.
            is_anonymous=is_anonymous,

            name=data.get("name") or None,
            surname=data.get("surname") or None,
            contact=data.get("contact") or None,
            email=data.get("email") or None,
            id_number=data.get("id_number") or None,
            nationality=data.get("nationality") or None,

            region=data.get("region") or None,
            location=data.get("location") or None,

            incident_date=data.get("incident_date") or None,
            incident_time=data.get("incident_time") or None,

            age_group=data.get("age_group") or None,
            gender=data.get("gender") or None,

            sector=data.get("sector") or "none",
            institution=data.get("institution") or None,

            amount=data.get("amount") or None,

            description=data.get("description") or "",

            evidence=request.FILES.get("evidence"),
        )

        return JsonResponse(
            {
                "success": True,
                "case_code": case.case_code,
                "message": "Report received successfully",
                "evidence_uploaded": bool(case.evidence),
                "username": account.username,
            },
            status=201
        )

    except IntegrityError:
        return JsonResponse(
            {
                "success": False,
                "error": (
                    "A case code conflict occurred. "
                    "Please submit the report again."
                )
            },
            status=409
        )

    except Exception as e:
        return JsonResponse(
            {
                "success": False,
                "error": str(e)
            },
            status=400
        )

@csrf_exempt
def get_my_reports(request):

    if request.method != "GET":
        return JsonResponse(
            {
                "success": False,
                "error": "Only GET requests are allowed"
            },
            status=405
        )

    # ============================================================
    # GET LOGGED-IN ACCOUNT
    # ============================================================

    username = request.GET.get("username", "").strip()

    if not username:
        return JsonResponse(
            {
                "success": False,
                "error": "Username is required"
            },
            status=400
        )

    # ============================================================
    # FIND ACCOUNT
    # ============================================================

    try:
        account = AnonymousAccount.objects.get(
            username=username
        )

    except AnonymousAccount.DoesNotExist:
        return JsonResponse(
            {
                "success": False,
                "error": "Account not found"
            },
            status=404
        )

    # ============================================================
    # GET ALL REPORTS SUBMITTED BY THIS ACCOUNT
    # ============================================================

    cases = Case.objects.filter(
        account=account
    ).order_by("-created_at")

    # ============================================================
    # BUILD RESPONSE
    # ============================================================

    reports = []

    for case in cases:

        reports.append(
            {
                "id": case.id,

                "case_code": case.case_code,

                # IMPORTANT:
                # This tells Flutter whether THIS report
                # was anonymous or not.
                "is_anonymous": case.is_anonymous,

                "id_number": case.id_number,

                "name": case.name,

                "surname": case.surname,

                "contact": case.contact,

                "email": case.email,

                "place": case.location,

                "date": (
                    case.incident_date.isoformat()
                    if case.incident_date
                    else None
                ),

                "time": (
                    case.incident_time.isoformat()
                    if case.incident_time
                    else None
                ),

                "department": "",

                "description": case.description,

                "attachments": (
                    [case.evidence.url]
                    if case.evidence
                    else []
                ),

                "status": case.status,

                "created_at": case.created_at.isoformat(),

                "nationality": case.nationality,

                "gender": case.gender,

                "age": case.age_group,

                "region": case.region,

                "sector": case.sector,

                "institution": case.institution,

                "amount_paid": (
                    str(case.amount)
                    if case.amount is not None
                    else None
                ),
            }
        )

    # ============================================================
    # RETURN REPORTS
    # ============================================================

    return JsonResponse(
        {
            "success": True,
            "reports": reports,
        }
    )
    
# ---------------- CASE DETAIL ----------------

def case_detail(request, id):

    case = get_object_or_404(
        Case,
        id=id
    )


    return render(
        request,
        "dashboard/case_detail.html",
        {
            "case": case
        }
    )



def update_case_status(request, id, status):

    case = get_object_or_404(
        Case,
        id=id
    )


    allowed_statuses = [
        "received",
        "investigation",
        "referral",
        "closed"
    ]


    if status in allowed_statuses:

        case.status = status

        case.save()
    AuditLog.objects.create(
    case=case,
    action="Status Updated",
    description=f"Status changed to {case.status}",
    user=request.user,
    ip_address=request.META.get("REMOTE_ADDR")
)

    return redirect(
    'case_detail',
    id=case.id
)

# ---------------- UPDATE STATUS ----------------

def update_status(request, id):

    case = get_object_or_404(
        Case,
        id=id
    )

    if request.method == "POST":

        status = request.POST.get("status")

        referral_notes = request.POST.get(
            "referral_notes"
        )

        decline_reason = request.POST.get(
            "decline_reason"
        )


        # Status remains the actual workflow status
        case.status = status


        # Referral keeps referral as status
        # Explanation is stored separately
        if status == "referral":

            case.referral_notes = referral_notes

            case.decline_reason = ""


        # Declined keeps declined as status
        # Reason is stored separately
        elif status == "declined":

            case.decline_reason = decline_reason

            case.referral_notes = ""


        else:

            case.referral_notes = ""

            case.decline_reason = ""


        case.save()

        messages.success(
        request,
        f"Case {case.case_code} status updated successfully to {case.get_status_display()}."
)


        AuditLog.objects.create(
            case=case,
            action="Status Updated",
            description=f"Status changed to {case.status}",
            user=request.user,
            ip_address=request.META.get("REMOTE_ADDR")
        )


        return redirect(
            "cases"
        )


    return render(
        request,
        "dashboard/update_status.html",
        {
            "case": case
        }
    )
# ---------------- ANALYTICS ----------------

def analytics(request):

    return render(
        request,
        "dashboard/analytics.html"
    )



from .models import Case, AuditLog, OrganisationSettings
from .forms import OrganisationSettingsForm


def settings(request):

    organisation = OrganisationSettings.objects.first()

    if organisation is None:
        organisation = OrganisationSettings.objects.create()

    form = OrganisationSettingsForm(instance=organisation)


    if request.method == "POST":

        form = OrganisationSettingsForm(
            request.POST,
            request.FILES,
            instance=organisation
        )

        if form.is_valid():
            form.save()


    return render(
        request,
        "dashboard/settings.html",
        {
            "form": form,
            "organisation": organisation,
        }
    )
def regions(request):

    regions_list = [
        "hhohho",
        "manzini",
        "lubombo",
        "shiselweni"
    ]


    hhohho = Case.objects.filter(
        region__iexact="hhohho"
    ).count()


    manzini = Case.objects.filter(
        region__iexact="manzini"
    ).count()


    lubombo = Case.objects.filter(
        region__iexact="lubombo"
    ).count()


    shiselweni = Case.objects.filter(
        region__iexact="shiselweni"
    ).count()



    total_cases = Case.objects.count()


    region_summary = []


    for region in regions_list:

        count = Case.objects.filter(
            region__iexact=region
        ).count()


        percentage = 0

        if total_cases > 0:
            percentage = round(
                (count / total_cases) * 100,
                1
            )


        region_summary.append({

            "name": region.title(),

            "count": count,

            "percentage": percentage

        })



    context = {

        "hhohho": hhohho,

        "manzini": manzini,

        "lubombo": lubombo,

        "shiselweni": shiselweni,


        "region_summary": region_summary,

    }
    print(region_summary)

    return render(
        request,
        "dashboard/regions.html",
        context
    )

def export_region_excel(request):

    from openpyxl import Workbook
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse

    cases = Case.objects.all()

    wb = Workbook()

    ws = wb.active
    ws.title = "Regional Analysis"


    # =========================================================
    # TITLE
    # =========================================================

    ws["A1"] = "ACC Regional Analysis"

    ws["A1"].font = Font(
        bold=True,
        size=16
    )

    ws.merge_cells("A1:C1")

    ws["A1"].alignment = Alignment(
        horizontal="center"
    )


    # =========================================================
    # HEADERS
    # =========================================================

    ws.append([])

    ws.append([
        "Region",
        "Number of Cases",
        "Percentage"
    ])


    header_row = 3


    for cell in ws[header_row]:

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="0A1F44"
        )

        cell.alignment = Alignment(
            horizontal="center"
        )


    # =========================================================
    # REGIONAL DATA
    # =========================================================

    total_cases = cases.count()


    regions = [
        "hhohho",
        "manzini",
        "lubombo",
        "shiselweni"
    ]


    for region in regions:

        count = cases.filter(
            region__iexact=region
        ).count()


        percentage = 0

        if total_cases > 0:

            percentage = round(
                (count / total_cases) * 100,
                2
            )


        ws.append([
            region.title(),
            count,
            percentage
        ])


    # =========================================================
    # FORMAT TABLE
    # =========================================================

    for row in range(4, 8):

        ws.cell(
            row=row,
            column=3
        ).number_format = '0.00"%"'


    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 18


    # =========================================================
    # BAR CHART
    # =========================================================

    bar_chart = BarChart()

    bar_chart.title = "Cases by Region"

    bar_chart.y_axis.title = "Number of Cases"

    bar_chart.x_axis.title = "Region"

    bar_chart.style = 10

    bar_chart.height = 8

    bar_chart.width = 15


    bar_data = Reference(
        ws,
        min_col=2,
        min_row=3,
        max_row=7
    )


    bar_categories = Reference(
        ws,
        min_col=1,
        min_row=4,
        max_row=7
    )


    bar_chart.add_data(
        bar_data,
        titles_from_data=True
    )

    bar_chart.set_categories(
        bar_categories
    )


    ws.add_chart(
        bar_chart,
        "E3"
    )


    # =========================================================
    # PIE CHART
    # =========================================================

    pie_chart = PieChart()

    pie_chart.title = "Regional Case Distribution"

    pie_chart.height = 8

    pie_chart.width = 12


    pie_data = Reference(
        ws,
        min_col=2,
        min_row=3,
        max_row=7
    )


    pie_labels = Reference(
        ws,
        min_col=1,
        min_row=4,
        max_row=7
    )


    pie_chart.add_data(
        pie_data,
        titles_from_data=True
    )

    pie_chart.set_categories(
        pie_labels
    )


    ws.add_chart(
        pie_chart,
        "E20"
    )


    # =========================================================
    # CREATE EXCEL RESPONSE
    # =========================================================

    response = HttpResponse(
        content_type=
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


    response["Content-Disposition"] = (
        'attachment; filename="ACC_Regional_Analysis.xlsx"'
    )


    wb.save(response)


    return response


def gender_analysis(request):

    male = Case.objects.filter(
        gender__iexact="male"
    ).count()

    female = Case.objects.filter(
        gender__iexact="female"
    ).count()


    total_gender = male + female


    male_percentage = round(
        (male / total_gender) * 100, 2
    ) if total_gender else 0


    female_percentage = round(
        (female / total_gender) * 100, 2
    ) if total_gender else 0


    return render(

        request,

        "dashboard/gender_analysis.html",

        {

            "male": male,

            "female": female,

            "male_percentage": male_percentage,

            "female_percentage": female_percentage,

        }

    )


def export_gender_excel(request):

    import openpyxl

    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.styles import Font, PatternFill, Alignment

    from django.http import HttpResponse

    from .models import Case


    # =========================================================
    # GET GENDER TOTALS
    # =========================================================

    male = Case.objects.filter(
        gender__iexact="male"
    ).count()


    female = Case.objects.filter(
        gender__iexact="female"
    ).count()


    total = male + female


    male_percentage = round(
        (male / total) * 100,
        2
    ) if total else 0


    female_percentage = round(
        (female / total) * 100,
        2
    ) if total else 0


    # =========================================================
    # CREATE WORKBOOK
    # =========================================================

    workbook = openpyxl.Workbook()

    sheet = workbook.active

    sheet.title = "Gender Analysis"


    # =========================================================
    # TITLE
    # =========================================================

    sheet["A1"] = "ACC Gender Analysis"

    sheet["A1"].font = Font(
        bold=True,
        size=16
    )

    sheet.merge_cells("A1:C1")

    sheet["A1"].alignment = Alignment(
        horizontal="center"
    )


    # =========================================================
    # HEADERS
    # =========================================================

    sheet.append([])

    sheet.append([
        "Gender",
        "Total Cases",
        "Percentage"
    ])


    header_row = 3


    for cell in sheet[header_row]:

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="0A1F44"
        )

        cell.alignment = Alignment(
            horizontal="center"
        )


    # =========================================================
    # DATA
    # =========================================================

    sheet.append([
        "Male",
        male,
        male_percentage
    ])


    sheet.append([
        "Female",
        female,
        female_percentage
    ])


    # =========================================================
    # FORMAT PERCENTAGES
    # =========================================================

    sheet["C4"].number_format = '0.00"%"'
    sheet["C5"].number_format = '0.00"%"'


    # =========================================================
    # COLUMN WIDTHS
    # =========================================================

    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 20
    sheet.column_dimensions["C"].width = 18


    # =========================================================
    # BAR CHART
    # =========================================================

    bar_chart = BarChart()

    bar_chart.title = "Cases by Gender"

    bar_chart.y_axis.title = "Number of Cases"

    bar_chart.x_axis.title = "Gender"

    bar_chart.style = 10

    bar_chart.height = 8

    bar_chart.width = 15


    bar_data = Reference(
        sheet,
        min_col=2,
        min_row=3,
        max_row=5
    )


    bar_categories = Reference(
        sheet,
        min_col=1,
        min_row=4,
        max_row=5
    )


    bar_chart.add_data(
        bar_data,
        titles_from_data=True
    )


    bar_chart.set_categories(
        bar_categories
    )


    sheet.add_chart(
        bar_chart,
        "E3"
    )


    # =========================================================
    # PIE CHART
    # =========================================================

    pie_chart = PieChart()

    pie_chart.title = "Gender Distribution"

    pie_chart.height = 8

    pie_chart.width = 12


    pie_data = Reference(
        sheet,
        min_col=2,
        min_row=3,
        max_row=5
    )


    pie_labels = Reference(
        sheet,
        min_col=1,
        min_row=4,
        max_row=5
    )


    pie_chart.add_data(
        pie_data,
        titles_from_data=True
    )


    pie_chart.set_categories(
        pie_labels
    )


    sheet.add_chart(
        pie_chart,
        "E20"
    )


    # =========================================================
    # DOWNLOAD EXCEL FILE
    # =========================================================

    response = HttpResponse(
        content_type=
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


    response["Content-Disposition"] = (
        'attachment; filename="ACC_Gender_Analysis.xlsx"'
    )


    workbook.save(response)


    return response


def public_institutions(request):

    ministry = Case.objects.filter(
        institution="ministry"
    ).count()

    department = Case.objects.filter(
        institution="department"
    ).count()

    municipality = Case.objects.filter(
        institution="municipality"
    ).count()

    parastatal = Case.objects.filter(
        institution="parastatal"
    ).count()

    police = Case.objects.filter(
        institution="police"
    ).count()

    public_health = Case.objects.filter(
        institution="public_health"
    ).count()

    public_school = Case.objects.filter(
        institution="public_school"
    ).count()

    public_university = Case.objects.filter(
        institution="public_university"
    ).count()

    procurement = Case.objects.filter(
        institution="procurement"
    ).count()

    revenue = Case.objects.filter(
        institution="revenue"
    ).count()

    other_public = Case.objects.filter(
        institution="other_public"
    ).count()


    # Institution not selected
    none_cases = Case.objects.filter(
        institution__isnull=True
    ).count()



    # TOTAL PUBLIC INSTITUTION CASES
    total_cases = sum([
        ministry,
        department,
        municipality,
        parastatal,
        police,
        public_health,
        public_school,
        public_university,
        procurement,
        revenue,
        other_public,
        none_cases
    ])



    # PERCENTAGE FUNCTION
    def calculate_percentage(value):

        if total_cases == 0:
            return 0

        return round((value / total_cases) * 100, 1)



    context = {

        # COUNTS
        "ministry": ministry,
        "department": department,
        "municipality": municipality,
        "parastatal": parastatal,
        "police": police,
        "public_health": public_health,
        "public_school": public_school,
        "public_university": public_university,
        "procurement": procurement,
        "revenue": revenue,
        "other_public": other_public,
        "none_cases": none_cases,


        # PERCENTAGES
        "ministry_percentage": calculate_percentage(ministry),
        "department_percentage": calculate_percentage(department),
        "municipality_percentage": calculate_percentage(municipality),
        "parastatal_percentage": calculate_percentage(parastatal),
        "police_percentage": calculate_percentage(police),
        "public_health_percentage": calculate_percentage(public_health),
        "public_school_percentage": calculate_percentage(public_school),
        "public_university_percentage": calculate_percentage(public_university),
        "procurement_percentage": calculate_percentage(procurement),
        "revenue_percentage": calculate_percentage(revenue),
        "other_public_percentage": calculate_percentage(other_public),
        "none_percentage": calculate_percentage(none_cases),

    }


    return render(
        request,
        "dashboard/public_institutions.html",
        context
    )


def export_public_institutions_excel(request):

    import openpyxl

    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.styles import Font, PatternFill, Alignment

    from django.http import HttpResponse

    from .models import Case


    # =========================================================
    # PUBLIC INSTITUTION DATA
    # =========================================================

    data = [

        ["Institution", "Cases", "Percentage"],

        [
            "Ministry",
            Case.objects.filter(
                institution="ministry"
            ).count(),
            ""
        ],

        [
            "Government Department",
            Case.objects.filter(
                institution="department"
            ).count(),
            ""
        ],

        [
            "Municipality",
            Case.objects.filter(
                institution="municipality"
            ).count(),
            ""
        ],

        [
            "Parastatal",
            Case.objects.filter(
                institution="parastatal"
            ).count(),
            ""
        ],

        [
            "Police",
            Case.objects.filter(
                institution="police"
            ).count(),
            ""
        ],

        [
            "Public Health Facility",
            Case.objects.filter(
                institution="public_health"
            ).count(),
            ""
        ],

        [
            "Public School",
            Case.objects.filter(
                institution="public_school"
            ).count(),
            ""
        ],

        [
            "Public University",
            Case.objects.filter(
                institution="public_university"
            ).count(),
            ""
        ],

        [
            "Procurement",
            Case.objects.filter(
                institution="procurement"
            ).count(),
            ""
        ],

        [
            "Revenue",
            Case.objects.filter(
                institution="revenue"
            ).count(),
            ""
        ],

        [
            "Other Public Institution",
            Case.objects.filter(
                institution="other_public"
            ).count(),
            ""
        ],

        [
            "None",
            Case.objects.filter(
                institution__isnull=True
            ).count(),
            ""
        ],
    ]


    # =========================================================
    # CALCULATE TOTAL
    # =========================================================

    total = sum(
        row[1]
        for row in data[1:]
    )


    # =========================================================
    # CALCULATE PERCENTAGES
    # =========================================================

    for row in data[1:]:

        if total > 0:

            row[2] = round(
                (row[1] / total) * 100,
                1
            )

        else:

            row[2] = 0


    # =========================================================
    # CREATE WORKBOOK
    # =========================================================

    workbook = openpyxl.Workbook()

    sheet = workbook.active

    sheet.title = "Public Institutions"


    # =========================================================
    # TITLE
    # =========================================================

    sheet["A1"] = "ACC Public Institutions Analysis"

    sheet["A1"].font = Font(
        bold=True,
        size=16
    )

    sheet.merge_cells("A1:C1")

    sheet["A1"].alignment = Alignment(
        horizontal="center"
    )


    # =========================================================
    # ADD DATA
    # =========================================================

    sheet.append([])


    for row in data:

        sheet.append(row)


    # Header is now row 3

    header_row = 3


    # =========================================================
    # STYLE HEADER
    # =========================================================

    for cell in sheet[header_row]:

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="0A1F44"
        )

        cell.alignment = Alignment(
            horizontal="center"
        )


    # =========================================================
    # FORMAT PERCENTAGES
    # =========================================================

    for row in range(4, 16):

        sheet.cell(
            row=row,
            column=3
        ).number_format = '0.0"%"'


    # =========================================================
    # COLUMN WIDTHS
    # =========================================================

    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 15
    sheet.column_dimensions["C"].width = 18


    # =========================================================
    # BAR CHART
    # =========================================================

    bar_chart = BarChart()

    bar_chart.title = "Cases by Public Institution"

    bar_chart.y_axis.title = "Number of Cases"

    bar_chart.x_axis.title = "Institution"

    bar_chart.style = 10

    bar_chart.height = 10

    bar_chart.width = 18


    bar_data = Reference(
        sheet,
        min_col=2,
        min_row=3,
        max_row=15
    )


    bar_categories = Reference(
        sheet,
        min_col=1,
        min_row=4,
        max_row=15
    )


    bar_chart.add_data(
        bar_data,
        titles_from_data=True
    )


    bar_chart.set_categories(
        bar_categories
    )


    sheet.add_chart(
        bar_chart,
        "E3"
    )


    # =========================================================
    # PIE CHART
    # =========================================================

    pie_chart = PieChart()

    pie_chart.title = "Public Institution Distribution"

    pie_chart.height = 10

    pie_chart.width = 15


    pie_data = Reference(
        sheet,
        min_col=2,
        min_row=3,
        max_row=15
    )


    pie_labels = Reference(
        sheet,
        min_col=1,
        min_row=4,
        max_row=15
    )


    pie_chart.add_data(
        pie_data,
        titles_from_data=True
    )


    pie_chart.set_categories(
        pie_labels
    )


    sheet.add_chart(
        pie_chart,
        "E24"
    )


    # =========================================================
    # CREATE RESPONSE
    # =========================================================

    response = HttpResponse(
        content_type=
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


    response["Content-Disposition"] = (
        'attachment; filename="ACC_public_institutions_analysis.xlsx"'
    )


    workbook.save(response)


    return response


def private_institutions(request):

    insurance = Case.objects.filter(
        institution="insurance"
    ).count()

    telecom = Case.objects.filter(
        institution="telecom"
    ).count()

    transport = Case.objects.filter(
        institution="transport"
    ).count()

    construction = Case.objects.filter(
        institution="construction"
    ).count()

    textile = Case.objects.filter(
        institution="textile"
    ).count()

    retail = Case.objects.filter(
        institution="retail"
    ).count()

    mining = Case.objects.filter(
        institution="mining"
    ).count()

    agriculture = Case.objects.filter(
        institution="agriculture"
    ).count()

    hotel = Case.objects.filter(
        institution="hotel"
    ).count()

    private_health = Case.objects.filter(
        institution="private_health"
    ).count()

    private_school = Case.objects.filter(
        institution="private_school"
    ).count()

    private_university = Case.objects.filter(
        institution="private_university"
    ).count()

    security = Case.objects.filter(
        institution="security"
    ).count()

    other_private = Case.objects.filter(
        institution="other_private"
    ).count()


    none_cases = Case.objects.filter(
        institution__isnull=True
    ).count()


    # TOTAL PRIVATE INSTITUTION CASES
    total_cases = sum([
        insurance,
        telecom,
        transport,
        construction,
        textile,
        retail,
        mining,
        agriculture,
        hotel,
        private_health,
        private_school,
        private_university,
        security,
        other_private,
        none_cases
    ])


    # PERCENTAGE FUNCTION
    def calculate_percentage(value):

        if total_cases == 0:
            return 0

        return round((value / total_cases) * 100, 1)



    context = {

        "insurance": insurance,
        "telecom": telecom,
        "transport": transport,
        "construction": construction,
        "textile": textile,
        "retail": retail,
        "mining": mining,
        "agriculture": agriculture,
        "hotel": hotel,
        "private_health": private_health,
        "private_school": private_school,
        "private_university": private_university,
        "security": security,
        "other_private": other_private,
        "none_cases": none_cases,


        # PERCENTAGES
        "insurance_percentage": calculate_percentage(insurance),
        "telecom_percentage": calculate_percentage(telecom),
        "transport_percentage": calculate_percentage(transport),
        "construction_percentage": calculate_percentage(construction),
        "textile_percentage": calculate_percentage(textile),
        "retail_percentage": calculate_percentage(retail),
        "mining_percentage": calculate_percentage(mining),
        "agriculture_percentage": calculate_percentage(agriculture),
        "hotel_percentage": calculate_percentage(hotel),
        "private_health_percentage": calculate_percentage(private_health),
        "private_school_percentage": calculate_percentage(private_school),
        "private_university_percentage": calculate_percentage(private_university),
        "security_percentage": calculate_percentage(security),
        "other_private_percentage": calculate_percentage(other_private),
        "none_percentage": calculate_percentage(none_cases),

    }


    return render(
        request,
        "dashboard/private_institutions.html",
        context
    )


def export_private_institutions_excel(request):

    import openpyxl

    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.styles import Font, PatternFill, Alignment

    from django.http import HttpResponse

    from .models import Case


    # =========================================================
    # GET PRIVATE SECTOR CASES
    # =========================================================

    cases = Case.objects.filter(
        sector="private"
    )


    # =========================================================
    # CREATE WORKBOOK
    # =========================================================

    workbook = openpyxl.Workbook()

    sheet = workbook.active

    sheet.title = "Private Institutions"


    # =========================================================
    # TITLE
    # =========================================================

    sheet["A1"] = "ACC Private Institutions Analysis"

    sheet["A1"].font = Font(
        bold=True,
        size=16
    )

    sheet.merge_cells("A1:C1")

    sheet["A1"].alignment = Alignment(
        horizontal="center"
    )


    # =========================================================
    # HEADERS
    # =========================================================

    sheet.append([])

    sheet.append([
        "Institution",
        "Cases",
        "Percentage"
    ])


    header_row = 3


    for cell in sheet[header_row]:

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="0A1F44"
        )

        cell.alignment = Alignment(
            horizontal="center"
        )


    # =========================================================
    # INSTITUTIONS
    # =========================================================

    institutions = [

        "bank",

        "private_school",

        "company",

        "other_private"

    ]


    total = cases.count()


    # =========================================================
    # ADD DATA
    # =========================================================

    for institution in institutions:

        count = cases.filter(
            institution=institution
        ).count()


        percentage = 0


        if total > 0:

            percentage = round(
                (count / total) * 100,
                1
            )


        sheet.append([

            institution.replace(
                "_",
                " "
            ).title(),

            count,

            percentage

        ])


    # =========================================================
    # FORMAT PERCENTAGES
    # =========================================================

    for row in range(4, 8):

        sheet.cell(
            row=row,
            column=3
        ).number_format = '0.0"%"'


    # =========================================================
    # COLUMN WIDTHS
    # =========================================================

    sheet.column_dimensions["A"].width = 30

    sheet.column_dimensions["B"].width = 15

    sheet.column_dimensions["C"].width = 18


    # =========================================================
    # BAR CHART
    # =========================================================

    bar_chart = BarChart()

    bar_chart.title = "Cases by Private Institution"

    bar_chart.y_axis.title = "Number of Cases"

    bar_chart.x_axis.title = "Institution"

    bar_chart.style = 10

    bar_chart.height = 8

    bar_chart.width = 16


    bar_data = Reference(
        sheet,
        min_col=2,
        min_row=3,
        max_row=7
    )


    bar_categories = Reference(
        sheet,
        min_col=1,
        min_row=4,
        max_row=7
    )


    bar_chart.add_data(
        bar_data,
        titles_from_data=True
    )


    bar_chart.set_categories(
        bar_categories
    )


    sheet.add_chart(
        bar_chart,
        "E3"
    )


    # =========================================================
    # PIE CHART
    # =========================================================

    pie_chart = PieChart()

    pie_chart.title = "Private Institution Distribution"

    pie_chart.height = 8

    pie_chart.width = 14


    pie_data = Reference(
        sheet,
        min_col=2,
        min_row=3,
        max_row=7
    )


    pie_labels = Reference(
        sheet,
        min_col=1,
        min_row=4,
        max_row=7
    )


    pie_chart.add_data(
        pie_data,
        titles_from_data=True
    )


    pie_chart.set_categories(
        pie_labels
    )


    sheet.add_chart(
        pie_chart,
        "E20"
    )


    # =========================================================
    # DOWNLOAD FILE
    # =========================================================

    response = HttpResponse(
        content_type=
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


    response["Content-Disposition"] = (
        'attachment; filename="ACC_Private_Institution_Analysis.xlsx"'
    )


    workbook.save(response)


    return response



def age_groups(request):

    under_18 = Case.objects.filter(
        age_group="under_18"
    ).count()


    age_18_24 = Case.objects.filter(
        age_group="18_24"
    ).count()


    age_25_34 = Case.objects.filter(
        age_group="25_34"
    ).count()


    age_35_44 = Case.objects.filter(
        age_group="35_44"
    ).count()


    age_45_54 = Case.objects.filter(
        age_group="45_54"
    ).count()


    age_55_64 = Case.objects.filter(
        age_group="55_64"
    ).count()


    above_65 = Case.objects.filter(
        age_group="above_65+"
    ).count()


    none_cases = Case.objects.filter(
        age_group__isnull=True
    ).count()


    total_age = (
        under_18 +
        age_18_24 +
        age_25_34 +
        age_35_44 +
        age_45_54 +
        age_55_64 +
        above_65 +
        none_cases
    )


    context = {

        "under_18": under_18,
        "age_18_24": age_18_24,
        "age_25_34": age_25_34,
        "age_35_44": age_35_44,
        "age_45_54": age_45_54,
        "age_55_64": age_55_64,
        "above_65": above_65,
        "none_cases": none_cases,


        "under_18_percentage": round((under_18 / total_age) * 100, 2) if total_age else 0,

        "age_18_24_percentage": round((age_18_24 / total_age) * 100, 2) if total_age else 0,

        "age_25_34_percentage": round((age_25_34 / total_age) * 100, 2) if total_age else 0,

        "age_35_44_percentage": round((age_35_44 / total_age) * 100, 2) if total_age else 0,

        "age_45_54_percentage": round((age_45_54 / total_age) * 100, 2) if total_age else 0,

        "age_55_64_percentage": round((age_55_64 / total_age) * 100, 2) if total_age else 0,

        "above_65_percentage": round((above_65 / total_age) * 100, 2) if total_age else 0,

        "none_cases_percentage": round((none_cases / total_age) * 100, 2) if total_age else 0,

    }


    return render(
        request,
        "dashboard/age_groups.html",
        context
    )

def export_age_groups_excel(request):

    import openpyxl

    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.styles import Font, PatternFill, Alignment

    from django.http import HttpResponse

    from .models import Case


    # =========================================================
    # GET AGE GROUP TOTALS
    # =========================================================

    under_18 = Case.objects.filter(
        age_group="under_18"
    ).count()


    age_18_24 = Case.objects.filter(
        age_group="18_24"
    ).count()


    age_25_34 = Case.objects.filter(
        age_group="25_34"
    ).count()


    age_35_44 = Case.objects.filter(
        age_group="35_44"
    ).count()


    age_45_54 = Case.objects.filter(
        age_group="45_54"
    ).count()


    age_55_64 = Case.objects.filter(
        age_group="55_64"
    ).count()


    above_65 = Case.objects.filter(
        age_group="above_65+"
    ).count()


    none_cases = Case.objects.filter(
        age_group__isnull=True
    ).count()


    # =========================================================
    # TOTAL
    # =========================================================

    total = (
        under_18
        + age_18_24
        + age_25_34
        + age_35_44
        + age_45_54
        + age_55_64
        + above_65
        + none_cases
    )


    # =========================================================
    # CREATE WORKBOOK
    # =========================================================

    workbook = openpyxl.Workbook()

    sheet = workbook.active

    sheet.title = "Age Groups"


    # =========================================================
    # TITLE
    # =========================================================

    sheet["A1"] = "ACC Age Group Analysis"

    sheet["A1"].font = Font(
        bold=True,
        size=16
    )

    sheet.merge_cells("A1:C1")

    sheet["A1"].alignment = Alignment(
        horizontal="center"
    )


    # =========================================================
    # HEADERS
    # =========================================================

    sheet.append([])

    sheet.append([
        "Age Group",
        "Cases",
        "Percentage"
    ])


    header_row = 3


    for cell in sheet[header_row]:

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="0A1F44"
        )

        cell.alignment = Alignment(
            horizontal="center"
        )


    # =========================================================
    # AGE GROUP DATA
    # =========================================================

    data = [

        ("Under 18", under_18),

        ("18 - 24", age_18_24),

        ("25 - 34", age_25_34),

        ("35 - 44", age_35_44),

        ("45 - 54", age_45_54),

        ("55 - 64", age_55_64),

        ("65+", above_65),

        ("None", none_cases),

    ]


    # =========================================================
    # ADD DATA TO WORKSHEET
    # =========================================================

    for label, count in data:

        percentage = round(
            (count / total) * 100,
            2
        ) if total else 0


        sheet.append([
            label,
            count,
            percentage
        ])


    # =========================================================
    # FORMAT PERCENTAGES
    # =========================================================

    for row in range(4, 12):

        sheet.cell(
            row=row,
            column=3
        ).number_format = '0.00"%"'


    # =========================================================
    # COLUMN WIDTHS
    # =========================================================

    sheet.column_dimensions["A"].width = 20

    sheet.column_dimensions["B"].width = 15

    sheet.column_dimensions["C"].width = 18


    # =========================================================
    # BAR CHART
    # =========================================================

    bar_chart = BarChart()

    bar_chart.title = "Cases by Age Group"

    bar_chart.y_axis.title = "Number of Cases"

    bar_chart.x_axis.title = "Age Group"

    bar_chart.style = 10

    bar_chart.height = 10

    bar_chart.width = 18


    bar_data = Reference(
        sheet,
        min_col=2,
        min_row=3,
        max_row=11
    )


    bar_categories = Reference(
        sheet,
        min_col=1,
        min_row=4,
        max_row=11
    )


    bar_chart.add_data(
        bar_data,
        titles_from_data=True
    )


    bar_chart.set_categories(
        bar_categories
    )


    sheet.add_chart(
        bar_chart,
        "E3"
    )


    # =========================================================
    # PIE CHART
    # =========================================================

    pie_chart = PieChart()

    pie_chart.title = "Age Group Distribution"

    pie_chart.height = 10

    pie_chart.width = 15


    pie_data = Reference(
        sheet,
        min_col=2,
        min_row=3,
        max_row=11
    )


    pie_labels = Reference(
        sheet,
        min_col=1,
        min_row=4,
        max_row=11
    )


    pie_chart.add_data(
        pie_data,
        titles_from_data=True
    )


    pie_chart.set_categories(
        pie_labels
    )


    sheet.add_chart(
        pie_chart,
        "E24"
    )


    # =========================================================
    # DOWNLOAD EXCEL FILE
    # =========================================================

    response = HttpResponse(
        content_type=
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


    response["Content-Disposition"] = (
        'attachment; filename="ACC_Age_Groups.xlsx"'
    )


    workbook.save(response)


    return response


from django.shortcuts import render
from django.http import HttpResponse
from .models import Case
import openpyxl


def get_reporter_type_data():

    total = Case.objects.count()

    anonymous = Case.objects.filter(
        is_anonymous=True
    ).count()

    identified = Case.objects.filter(
        is_anonymous=False
    ).count()


    anonymous_percentage = (
        anonymous / total * 100
    ) if total else 0


    identified_percentage = (
        identified / total * 100
    ) if total else 0


    return {
        "total": total,
        "anonymous": anonymous,
        "identified": identified,
        "anonymous_percentage": round(anonymous_percentage, 1),
        "identified_percentage": round(identified_percentage, 1),
    }



def reporter_type_analysis(request):

    total = Case.objects.count()


    anonymous = Case.objects.filter(
        is_anonymous=True
    ).count()


    identified = Case.objects.filter(
        is_anonymous=False
    ).count()



    anonymous_percentage = (
        anonymous / total * 100
    ) if total else 0



    identified_percentage = (
        identified / total * 100
    ) if total else 0



    return render(
        request,
        "dashboard/reporter_type_analysis.html",
        {

            "total": total,

            "anonymous": anonymous,

            "identified": identified,

            "anonymous_percentage": round(
                anonymous_percentage, 1
            ),

            "identified_percentage": round(
                identified_percentage, 1
            ),


            "reporter_labels": json.dumps([
                "Anonymous",
                "Revealed"
            ]),


            "reporter_values": json.dumps([
                anonymous,
                identified
            ]),

        },
    )
import openpyxl
from django.http import HttpResponse


def export_reporter_type_excel(request):

    data = get_reporter_type_data()


    workbook = openpyxl.Workbook()

    sheet = workbook.active

    sheet.title = "Reporter Type Analysis"


    sheet.append([
        "Reporter Type",
        "Number of Cases",
        "Percentage"
    ])


    sheet.append([
        "Anonymous",
        data["anonymous"],
        f'{data["anonymous_percentage"]}%'
    ])


    sheet.append([
        "Identified",
        data["identified"],
        f'{data["identified_percentage"]}%'
    ])


    sheet.append([
        "Total Cases",
        data["total"],
        "100%"
    ])


    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


    response["Content-Disposition"] = (
        "attachment; filename=reporter_type_analysis.xlsx"
    )


    workbook.save(response)


    return response


from django.db.models import Count
from django.db.models.functions import TruncMonth
from django.shortcuts import render
from dashboard.models import Case
import json


def trends(request):

    monthly_cases = (
        Case.objects
        .annotate(month=TruncMonth("created_at"))
        .values("month")
        .annotate(total=Count("id"))
        .order_by("month")
    )


    monthly_labels = []

    monthly_values = []


    for item in monthly_cases:

        monthly_labels.append(
            item["month"].strftime("%b %Y")
        )

        monthly_values.append(
            item["total"]
        )


    public_total = Case.objects.filter(
        sector="public"
    ).count()


    private_total = Case.objects.filter(
        sector="private"
    ).count()


    none_total = (
        Case.objects.filter(
            sector__isnull=True
        ).count()
        +
        Case.objects.filter(
            sector="none"
        ).count()
    )


    received = Case.objects.filter(
        status="received"
    ).count()


    investigation = Case.objects.filter(
        status="investigation"
    ).count()


    referral = Case.objects.filter(
        status="referral"
    ).count()


    declined = Case.objects.filter(
        status="declined"
    ).count()


    closed = Case.objects.filter(
        status="closed"
    ).count()



    total_cases = Case.objects.count()



    context = {


        "monthly_labels": json.dumps(
            monthly_labels
        ),


        "monthly_values": json.dumps(
            monthly_values
        ),



        "public_total": public_total,

        "private_total": private_total,

        "none_total": none_total,



        "public_percentage": round(
            (public_total / total_cases) * 100, 2
        ) if total_cases else 0,


        "private_percentage": round(
            (private_total / total_cases) * 100, 2
        ) if total_cases else 0,


        "none_percentage": round(
            (none_total / total_cases) * 100, 2
        ) if total_cases else 0,



        "status_labels": json.dumps(
            [
                "Received",
                "Under Investigation",
                "Referral",
                "Declined",
                "Closed"
            ]
        ),



        # Status totals

        "received": received,

        "investigation": investigation,

        "referral": referral,

        "declined": declined,

        "closed": closed,



        # Chart.js data

        "received_values": json.dumps(
            [received]
        ),


        "investigation_values": json.dumps(
            [investigation]
        ),


        "referral_values": json.dumps(
            [referral]
        ),


        "declined_values": json.dumps(
            [declined]
        ),


        "closed_values": json.dumps(
            [closed]
        ),



        # Percentages

        "received_percentage": round(
            (received / total_cases) * 100, 2
        ) if total_cases else 0,


        "investigation_percentage": round(
            (investigation / total_cases) * 100, 2
        ) if total_cases else 0,


        "referral_percentage": round(
            (referral / total_cases) * 100, 2
        ) if total_cases else 0,


        "declined_percentage": round(
            (declined / total_cases) * 100, 2
        ) if total_cases else 0,


        "closed_percentage": round(
            (closed / total_cases) * 100, 2
        ) if total_cases else 0,

    }


    return render(
        request,
        "dashboard/trends.html",
        context
    )
    


def export_trends_excel(request):

    import openpyxl

    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.styles import Font, PatternFill, Alignment

    from django.http import HttpResponse

    from .models import Case


    # =========================================================
    # SECTOR TOTALS
    # =========================================================

    public_total = Case.objects.filter(
        sector="public"
    ).count()


    private_total = Case.objects.filter(
        sector="private"
    ).count()


    # =========================================================
    # STATUS TOTALS
    # =========================================================

    received = Case.objects.filter(
        status="received"
    ).count()


    investigation = Case.objects.filter(
        status="investigation"
    ).count()


    referral = Case.objects.filter(
        status="referral"
    ).count()


    declined = Case.objects.filter(
        status="declined"
    ).count()


    closed = Case.objects.filter(
        status="closed"
    ).count()


    # =========================================================
    # TOTAL CASES
    # =========================================================

    total_cases = Case.objects.count()


    # =========================================================
    # CREATE WORKBOOK
    # =========================================================

    workbook = openpyxl.Workbook()

    sheet = workbook.active

    sheet.title = "Trends Analysis"


    # =========================================================
    # TITLE
    # =========================================================

    sheet["A1"] = "ACC Trends Analysis"

    sheet["A1"].font = Font(
        bold=True,
        size=16
    )

    sheet.merge_cells("A1:C1")

    sheet["A1"].alignment = Alignment(
        horizontal="center"
    )


    # =========================================================
    # HEADERS
    # =========================================================

    sheet.append([])

    sheet.append([
        "Category",
        "Cases",
        "Percentage"
    ])


    header_row = 3


    for cell in sheet[header_row]:

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="0A1F44"
        )

        cell.alignment = Alignment(
            horizontal="center"
        )


    # =========================================================
    # DATA
    # =========================================================

    data = [

        ("Public Sector", public_total),

        ("Private Sector", private_total),

        ("Received", received),

        ("Under Investigation", investigation),

        ("Referral", referral),

        ("Declined", declined),

        ("Closed", closed),

    ]


    # =========================================================
    # ADD DATA
    # =========================================================

    for label, count in data:

        percentage = round(
            (count / total_cases) * 100,
            2
        ) if total_cases else 0


        sheet.append([
            label,
            count,
            percentage
        ])


    # =========================================================
    # FORMAT PERCENTAGES
    # =========================================================

    for row in range(4, 11):

        sheet.cell(
            row=row,
            column=3
        ).number_format = '0.00"%"'


    # =========================================================
    # COLUMN WIDTHS
    # =========================================================

    sheet.column_dimensions["A"].width = 28

    sheet.column_dimensions["B"].width = 15

    sheet.column_dimensions["C"].width = 18


    # =========================================================
    # BAR CHART — ALL CATEGORIES
    # =========================================================

    bar_chart = BarChart()

    bar_chart.title = "Cases by Category"

    bar_chart.y_axis.title = "Number of Cases"

    bar_chart.x_axis.title = "Category"

    bar_chart.style = 10

    bar_chart.height = 10

    bar_chart.width = 18


    bar_data = Reference(
        sheet,
        min_col=2,
        min_row=3,
        max_row=10
    )


    bar_categories = Reference(
        sheet,
        min_col=1,
        min_row=4,
        max_row=10
    )


    bar_chart.add_data(
        bar_data,
        titles_from_data=True
    )


    bar_chart.set_categories(
        bar_categories
    )


    sheet.add_chart(
        bar_chart,
        "E3"
    )


    # =========================================================
    # SECTOR PIE CHART
    # =========================================================

    pie_sector = PieChart()

    pie_sector.title = "Sector Distribution"

    pie_sector.height = 8

    pie_sector.width = 14


    # Public and Private are rows 4 and 5

    sector_data = Reference(
        sheet,
        min_col=2,
        min_row=3,
        max_row=5
    )


    sector_labels = Reference(
        sheet,
        min_col=1,
        min_row=4,
        max_row=5
    )


    pie_sector.add_data(
        sector_data,
        titles_from_data=True
    )


    pie_sector.set_categories(
        sector_labels
    )


    sheet.add_chart(
        pie_sector,
        "E24"
    )


    # =========================================================
    # STATUS PIE CHART
    # =========================================================

    pie_status = PieChart()

    pie_status.title = "Status Distribution"

    pie_status.height = 8

    pie_status.width = 14


    # Status rows:
    # Received = 6
    # Investigation = 7
    # Referral = 8
    # Declined = 9
    # Closed = 10

    status_data = Reference(
        sheet,
        min_col=2,
        min_row=3,
        max_row=10
    )


    status_labels = Reference(
        sheet,
        min_col=1,
        min_row=4,
        max_row=10
    )


    # We need only the status rows.
    # Therefore create a separate status data section.

    sheet["E35"] = "Status"

    sheet["F35"] = "Cases"


    sheet["E36"] = "Received"
    sheet["F36"] = received

    sheet["E37"] = "Under Investigation"
    sheet["F37"] = investigation

    sheet["E38"] = "Referral"
    sheet["F38"] = referral

    sheet["E39"] = "Declined"
    sheet["F39"] = declined

    sheet["E40"] = "Closed"
    sheet["F40"] = closed


    status_chart_data = Reference(
        sheet,
        min_col=6,
        min_row=35,
        max_row=40
    )


    status_chart_labels = Reference(
        sheet,
        min_col=5,
        min_row=36,
        max_row=40
    )


    pie_status.add_data(
        status_chart_data,
        titles_from_data=True
    )


    pie_status.set_categories(
        status_chart_labels
    )


    sheet.add_chart(
        pie_status,
        "E42"
    )


    # =========================================================
    # DOWNLOAD EXCEL FILE
    # =========================================================

    response = HttpResponse(
        content_type=
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


    response["Content-Disposition"] = (
        'attachment; filename="ACC_Trends_Analysis.xlsx"'
    )


    workbook.save(response)


    return response



def add_notification(request):

    if request.method == "POST":

        title = request.POST.get("title")
        description = request.POST.get("description")
        date = request.POST.get("date")
        time = request.POST.get("time")


        Notification.objects.create(
            title=title,
            description=description,
            date=date,
            time=time
        )


        return redirect("dashboard")


    return render(
        request,
        "dashboard/add_notification.html"
    )
from django.shortcuts import render
from django.db.models import Count
from django.db.models.functions import TruncMonth
from django.utils import timezone
import json

from cases.models import Report
from dashboard.models import Case


def overview(request):

    # ======================
    # SUMMARY CARDS
    # ======================

    total_cases = Case.objects.count()

    investigation = Case.objects.filter(
        status="investigation"
    ).count()

    referrals = Case.objects.filter(
        status="referral"
    ).count()

    closed = Case.objects.filter(
        status="closed"
    ).count()


    # ======================
    # SECTOR GRAPH
    # ======================

    
    public_total = Case.objects.filter(
        sector="public"
    ).count()


    private_total = Case.objects.filter(
        sector="private"
    ).count()


    none_total = (
        Case.objects.filter(
            sector__isnull=True
        ).count()
        +
        Case.objects.filter(
            sector="none"
        ).count()
    )



    # ======================
    # REGION GRAPH
    # ======================

    regions = (
        Case.objects
        .values("region")
        .annotate(total=Count("id"))
    )

    region_labels = []
    region_values = []

    for item in regions:
        region_labels.append(item["region"])
        region_values.append(item["total"])



    # ======================
    # STATUS GRAPH
    # ======================

    statuses = (
        Case.objects
        .values("status")
        .annotate(total=Count("id"))
    )

    status_labels = []
    status_values = []

    for item in statuses:
        status_labels.append(item["status"])
        status_values.append(item["total"])



    # ======================
    # GENDER GRAPH
    # ======================

    genders = (
        Case.objects
        .values("gender")
        .annotate(total=Count("id"))
    )

    gender_labels = []
    gender_values = []

    for item in genders:
        gender_labels.append(item["gender"])
        gender_values.append(item["total"])



    # ======================
    # AGE GROUP GRAPH
    # ======================

    age_groups = (
        Case.objects
        .values("age_group")
        .annotate(total=Count("id"))
    )

    age_labels = []
    age_values = []

    for item in age_groups:
        age_labels.append(item["age_group"])
        age_values.append(item["total"])



    # ======================
    # MONTHLY TREND
    # ======================

    monthly_cases = (
        Case.objects
        .annotate(month=TruncMonth("created_at"))
        .values("month")
        .annotate(total=Count("id"))
        .order_by("month")
    )


    monthly_labels = []
    monthly_values = []


    for item in monthly_cases:

        monthly_labels.append(
            item["month"].strftime("%b %Y")
        )

        monthly_values.append(
            item["total"]
        )



    context = {


        # CARDS

        "total_cases": total_cases,

        "investigation": investigation,

        "referrals": referrals,

        "closed": closed,



        # SECTOR

        "public_total": public_total,
        
                "private_total": private_total,
        
                "none_total": none_total,
        
        
        
                "public_percentage": round(
                    (public_total / total_cases) * 100, 2
                ) if total_cases else 0,
        
        
                "private_percentage": round(
                    (private_total / total_cases) * 100, 2
                ) if total_cases else 0,
        
        
                "none_percentage": round(
                    (none_total / total_cases) * 100, 2
                ) if total_cases else 0,
        


        # REGION

        "region_labels": json.dumps(region_labels),

        "region_values": json.dumps(region_values),



        # STATUS

        "status_labels": json.dumps(status_labels),

        "status_values": json.dumps(status_values),



        # GENDER

        "gender_labels": json.dumps(gender_labels),

        "gender_values": json.dumps(gender_values),



        # AGE

        "age_labels": json.dumps(age_labels),

        "age_values": json.dumps(age_values),



        # MONTHLY

        "monthly_labels": json.dumps(monthly_labels),

        "monthly_values": json.dumps(monthly_values),


    }


    return render(
        request,
        "dashboard/overview.html",
        context
    )

def export_overview_excel(request):

    import openpyxl

    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.styles import Font, PatternFill, Alignment

    from django.http import HttpResponse

    from .models import Case


    # =========================================================
    # GET ALL CASES
    # =========================================================

    cases = Case.objects.all()

    total_cases = cases.count()


    # =========================================================
    # CREATE WORKBOOK
    # =========================================================

    workbook = openpyxl.Workbook()

    sheet = workbook.active

    sheet.title = "Analytics Overview"


    # =========================================================
    # TITLE
    # =========================================================

    sheet["A1"] = "ACC Analytics Overview"

    sheet["A1"].font = Font(
        bold=True,
        size=16
    )

    sheet.merge_cells("A1:B1")

    sheet["A1"].alignment = Alignment(
        horizontal="center"
    )


    # =========================================================
    # SUMMARY HEADERS
    # =========================================================

    sheet.append([])

    sheet.append([
        "Metric",
        "Value"
    ])


    for cell in sheet[3]:

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="0A1F44"
        )

        cell.alignment = Alignment(
            horizontal="center"
        )


    # =========================================================
    # SUMMARY DATA
    # =========================================================

    summary_data = [

        (
            "Total Cases",
            total_cases
        ),

        (
            "Received",
            cases.filter(
                status="received"
            ).count()
        ),

        (
            "Under Investigation",
            cases.filter(
                status="investigation"
            ).count()
        ),

        (
            "Referral Cases",
            cases.filter(
                status="referral"
            ).count()
        ),

        (
            "Declined Cases",
            cases.filter(
                status="declined"
            ).count()
        ),

        (
            "Closed Cases",
            cases.filter(
                status="closed"
            ).count()
        ),

    ]


    for label, value in summary_data:

        sheet.append([
            label,
            value
        ])


    sheet.column_dimensions["A"].width = 28

    sheet.column_dimensions["B"].width = 15


    # =========================================================
    # REGIONAL ANALYSIS
    # =========================================================

    region_sheet = workbook.create_sheet(
        title="Regional Analysis"
    )


    region_sheet.append([
        "Region",
        "Cases"
    ])


    regions = [
        "hhohho",
        "manzini",
        "lubombo",
        "shiselweni"
    ]


    for region in regions:

        region_sheet.append([

            region.title(),

            cases.filter(
                region__iexact=region
            ).count()

        ])


    # Style header

    for cell in region_sheet[1]:

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="0A1F44"
        )


    region_sheet.column_dimensions["A"].width = 20
    region_sheet.column_dimensions["B"].width = 15


    # Regional chart

    region_chart = BarChart()

    region_chart.title = "Cases by Region"

    region_chart.y_axis.title = "Number of Cases"

    region_chart.x_axis.title = "Region"

    region_chart.height = 8

    region_chart.width = 14


    region_data = Reference(
        region_sheet,
        min_col=2,
        min_row=1,
        max_row=5
    )


    region_labels = Reference(
        region_sheet,
        min_col=1,
        min_row=2,
        max_row=5
    )


    region_chart.add_data(
        region_data,
        titles_from_data=True
    )

    region_chart.set_categories(
        region_labels
    )


    region_sheet.add_chart(
        region_chart,
        "D2"
    )


    # =========================================================
    # GENDER ANALYSIS
    # =========================================================

    gender_sheet = workbook.create_sheet(
        title="Gender Analysis"
    )


    gender_sheet.append([
        "Gender",
        "Cases"
    ])


    for gender in [
        "male",
        "female"
    ]:

        gender_sheet.append([

            gender.title(),

            cases.filter(
                gender__iexact=gender
            ).count()

        ])


    for cell in gender_sheet[1]:

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="0A1F44"
        )


    gender_sheet.column_dimensions["A"].width = 20
    gender_sheet.column_dimensions["B"].width = 15


    # Gender chart

    gender_chart = PieChart()

    gender_chart.title = "Gender Distribution"

    gender_chart.height = 8

    gender_chart.width = 12


    gender_data = Reference(
        gender_sheet,
        min_col=2,
        min_row=1,
        max_row=3
    )


    gender_labels = Reference(
        gender_sheet,
        min_col=1,
        min_row=2,
        max_row=3
    )


    gender_chart.add_data(
        gender_data,
        titles_from_data=True
    )

    gender_chart.set_categories(
        gender_labels
    )


    gender_sheet.add_chart(
        gender_chart,
        "D2"
    )


    # =========================================================
    # SECTOR ANALYSIS
    # =========================================================

    sector_sheet = workbook.create_sheet(
        title="Sector Analysis"
    )


    sector_sheet.append([
        "Sector",
        "Cases"
    ])


    for sector in [
        "public",
        "private"
    ]:

        sector_sheet.append([

            sector.title(),

            cases.filter(
                sector=sector
            ).count()

        ])


    for cell in sector_sheet[1]:

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="0A1F44"
        )


    sector_sheet.column_dimensions["A"].width = 20
    sector_sheet.column_dimensions["B"].width = 15


    # Sector chart

    sector_chart = PieChart()

    sector_chart.title = "Sector Distribution"

    sector_chart.height = 8

    sector_chart.width = 12


    sector_data = Reference(
        sector_sheet,
        min_col=2,
        min_row=1,
        max_row=3
    )


    sector_labels = Reference(
        sector_sheet,
        min_col=1,
        min_row=2,
        max_row=3
    )


    sector_chart.add_data(
        sector_data,
        titles_from_data=True
    )

    sector_chart.set_categories(
        sector_labels
    )


    sector_sheet.add_chart(
        sector_chart,
        "D2"
    )


    # =========================================================
    # STATUS ANALYSIS
    # =========================================================

    status_sheet = workbook.create_sheet(
        title="Status Analysis"
    )


    status_sheet.append([
        "Status",
        "Cases"
    ])


    statuses = [

        ("Received", "received"),

        ("Under Investigation", "investigation"),

        ("Referral", "referral"),

        ("Declined", "declined"),

        ("Closed", "closed"),

    ]


    for label, status_value in statuses:

        status_sheet.append([

            label,

            cases.filter(
                status=status_value
            ).count()

        ])


    for cell in status_sheet[1]:

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="0A1F44"
        )


    status_sheet.column_dimensions["A"].width = 28
    status_sheet.column_dimensions["B"].width = 15


    # Status chart

    status_chart = BarChart()

    status_chart.title = "Cases by Status"

    status_chart.y_axis.title = "Number of Cases"

    status_chart.x_axis.title = "Status"

    status_chart.height = 8

    status_chart.width = 16


    status_data = Reference(
        status_sheet,
        min_col=2,
        min_row=1,
        max_row=6
    )


    status_labels = Reference(
        status_sheet,
        min_col=1,
        min_row=2,
        max_row=6
    )


    status_chart.add_data(
        status_data,
        titles_from_data=True
    )

    status_chart.set_categories(
        status_labels
    )


    status_sheet.add_chart(
        status_chart,
        "D2"
    )


    # =========================================================
    # CREATE DOWNLOAD RESPONSE
    # =========================================================

    response = HttpResponse(
        content_type=
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


    response["Content-Disposition"] = (
        'attachment; filename="ACC_Analytics_Overview.xlsx"'
    )


    workbook.save(response)


    return response


from django.contrib.auth.views import PasswordResetConfirmView
from .models import AuditLog


class ACCPasswordResetConfirmView(PasswordResetConfirmView):

    def form_valid(self, form):

        response = super().form_valid(form)

        AuditLog.objects.create(
            user=self.user,
            action="Password Changed",
            description=f"{self.user.username} successfully changed their password"
        )

        return response

from .models import CaseManagementSettings
from .forms import CaseManagementSettingsForm
def case_management_settings(request):

    settings = CaseManagementSettings.objects.first()

    if settings is None:
        settings = CaseManagementSettings.objects.create()

    if request.method == "POST":

        form = CaseManagementSettingsForm(
            request.POST,
            instance=settings
        )

        if form.is_valid():
            form.save()

            return redirect(
                "case_management_settings"
            )

    else:

        form = CaseManagementSettingsForm(
            instance=settings
        )


    return render(
        request,
        "dashboard/case_management_settings.html",
        {
            "form": form
        }
    )
from django.contrib.auth.hashers import make_password
from .models import AnonymousAccount

from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
def register_anonymous_account(request):

    if request.method != "POST":
        return JsonResponse(
            {
                "success": False,
                "error": "Only POST requests are allowed"
            },
            status=405
        )

    try:
        data = json.loads(request.body)

        # Normalize username so accidental spaces/capitalization
        # do not create what looks like a duplicate account.
        username = data.get("username", "").strip().lower()
        password = data.get("password", "")
        first_school = data.get("first_school", "").strip()
        favourite_month = data.get("favourite_month", "").strip()

        # ================= REQUIRED FIELDS =================
        if not username or not password or not first_school or not favourite_month:
            return JsonResponse(
                {
                    "success": False,
                    "error": "All fields are required"
                },
                status=400
            )

        # ================= DUPLICATE USERNAME CHECK =================
        if AnonymousAccount.objects.filter(
            username=username
        ).exists():

            return JsonResponse(
                {
                    "success": False,
                    "error": (
                        "This anonymous username already exists. "
                        "Please choose a different username."
                    )
                },
                status=409
            )

        # ================= CREATE ACCOUNT =================
        account = AnonymousAccount.objects.create(
            username=username,
            password=make_password(password),
            first_school=first_school,
            favourite_month=favourite_month
        )

        return JsonResponse(
            {
                "success": True,
                "message": "Anonymous account created successfully",
                "username": account.username
            },
            status=201
        )

    except Exception as e:

        return JsonResponse(
            {
                "success": False,
                "error": str(e)
            },
            status=400
        )
from django.contrib.auth.hashers import make_password, check_password    
from rest_framework_simplejwt.tokens import RefreshToken

@csrf_exempt
@csrf_exempt
def login_anonymous_account(request):
    if request.method != "POST":
        return JsonResponse(
            {
                "success": False,
                "error": "Only POST requests are allowed",
            },
            status=405,
        )

    try:
        data = json.loads(request.body)

        username = data.get("username", "").strip()
        password = data.get("password", "")

        if not username or not password:
            return JsonResponse(
                {
                    "success": False,
                    "error": "Username and password are required",
                },
                status=400,
            )

        try:
            account = AnonymousAccount.objects.get(
                username=username
            )
        except AnonymousAccount.DoesNotExist:
            return JsonResponse(
                {
                    "success": False,
                    "error": "Invalid username or password",
                },
                status=401,
            )

        if not check_password(
            password,
            account.password,
        ):
            return JsonResponse(
                {
                    "success": False,
                    "error": "Invalid username or password",
                },
                status=401,
            )

        # Create JWT refresh token
        refresh = RefreshToken.for_user(account)

        # Add anonymous username to the tokens
        refresh["username"] = account.username

        access = refresh.access_token
        access["username"] = account.username

        return JsonResponse(
            {
                "success": True,
                "message": "Login successful",
                "username": account.username,
                "access": str(access),
                "refresh": str(refresh),
            },
            status=200,
        )

    except json.JSONDecodeError:
        return JsonResponse(
            {
                "success": False,
                "error": "Invalid JSON data",
            },
            status=400,
        )

    except Exception as e:
        return JsonResponse(
            {
                "success": False,
                "error": str(e),
            },
            status=500,
        )

from django.contrib.auth.hashers import make_password
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
import json

from .models import AnonymousAccount


@csrf_exempt
def reset_anonymous_password(request):

    if request.method != "POST":
        return JsonResponse(
            {
                "success": False,
                "error": "Only POST requests are allowed"
            },
            status=405
        )

    try:
        data = json.loads(request.body)

        username = data.get("username", "").strip()
        first_school = data.get("first_school", "").strip()
        favourite_month = data.get("favourite_month", "").strip()
        new_password = data.get("new_password", "")

        if (
            not username
            or not first_school
            or not favourite_month
            or not new_password
        ):
            return JsonResponse(
                {
                    "success": False,
                    "error": "All fields are required"
                },
                status=400
            )

        try:
            account = AnonymousAccount.objects.get(
                username=username
            )

        except AnonymousAccount.DoesNotExist:
            return JsonResponse(
                {
                    "success": False,
                    "error": "Invalid username or recovery information"
                },
                status=401
            )

        # Check recovery answers
        if (
            account.first_school.strip().lower()
            != first_school.lower()
            or
            account.favourite_month.strip().lower()
            != favourite_month.lower()
        ):
            return JsonResponse(
                {
                    "success": False,
                    "error": "Invalid username or recovery information"
                },
                status=401
            )

        # Save the NEW password securely hashed
        account.password = make_password(new_password)
        account.save(update_fields=["password"])

        return JsonResponse(
            {
                "success": True,
                "message": "Password reset successful"
            },
            status=200
        )

    except json.JSONDecodeError:
        return JsonResponse(
            {
                "success": False,
                "error": "Invalid JSON data"
            },
            status=400
        )

    except Exception as e:
        return JsonResponse(
            {
                "success": False,
                "error": str(e)
            },
            status=500
        )

